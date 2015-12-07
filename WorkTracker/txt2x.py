#!/usr/bin/python

import os
import openpyxl
#from openpyxl.styles import Font, Style
                
def txt2x(job_num, item, pcmk, welder, plate, pipe, now):           
    os.chdir("/home/yojimbo/Documents/")

    def write_xlsx():
        # this creates the headers for the xlsx file
        a = ws.cell('A1')   
        a.value = "Job Number %s" % job_num    
        b = ws.cell('B1')
        b.value = "Item %s" % item
        a = ws.cell('A2')
        a.value = "PCMK"
        b = ws.cell('B2')
        b.value = "Welder ID"
        c = ws.cell('C2')
        c.value = "Plate Heat Number"
        d = ws.cell('D2')
        d.value = "Pipe Heat Number"
        e = ws.cell('E2')
        e.value = "Datetime"

        row = int(pcmk)+2

        write_to_cell_A = ws.cell('A%d' % row)
        write_to_cell_A.value = "%s" % pcmk
        write_to_cell_B = ws.cell('B%d' % row)
        write_to_cell_B.value = "%s" % welder
        write_to_cell_C = ws.cell('C%d' % row)
        write_to_cell_C.value = "%s" % plate
        write_to_cell_D = ws.cell('D%s' % row)
        write_to_cell_D.value = "%s" % pipe
        write_to_cell_E = ws.cell('E%d' % row)
        write_to_cell_E.value = "%s" % now

    # try to open a xlsx file by job number if sucessful 
    # open it and write to file 
    try:
        print "file existed"
        wb = openpyxl.load_workbook('%sitem%s.xlsx' % (job_num, item))
        ws = wb.get_sheet_by_name('Sheet')
        write_xlsx()
    # If tying to open file fails create the file and write to it
    except:
        print "created a file"
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(0)
        write_xlsx()

        # save all the work that was just done
    
    wb.save('%sitem%s.xlsx' % (job_num, item))
     
